#!/usr/bin/env python3
"""One-time extraction: build the sentiment base series from input/情绪指标.xlsx.

The Excel ships the full daily history of every raw input the six sentiment
indicators need. Extracting it locally means zero API calls for history;
scripts/update_sentiment_index.py only fetches increments after the Excel's
last date.

Output: data/raw/sentiment_base.csv
    date, pe_ttm, yield_10y, free_turn, amt, pct_chg, fund_eq, fund_all, close
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "input" / "情绪指标.xlsx"
OUT = ROOT / "data" / "raw" / "sentiment_base.csv"


def sheet_df(wb, name: str, min_row: int, cols: list[str], date_col: str = "Date") -> pd.DataFrame:
    ws = wb[name]
    rows = [r[: len(cols)] for r in ws.iter_rows(min_row=min_row, values_only=True) if r[0] is not None]
    df = pd.DataFrame(rows, columns=cols)
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=[date_col]).set_index(date_col)
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df.sort_index()


def main() -> None:
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    pe = sheet_df(wb, "1、股债收益差", 7, ["Date", "pe_ttm", "yield_10y"])
    turn = sheet_df(wb, "2、沪指自由流通换手率", 7, ["Date", "free_turn", "ma20", "p3", "p5"])[["free_turn"]]
    illiq = sheet_df(wb, "3、流动性冲击", 7, ["Date", "amt", "pct_chg", "illiq", "ma20", "ma60", "ls", "p3", "p5"])[["amt", "pct_chg"]]
    fund = sheet_df(wb, "4、30日新发基金发行占比", 3, ["Date", "fund_eq", "fund_all", "ratio", "p3", "p5"])[["fund_eq", "fund_all"]]

    # 收盘价在「等权情绪指标」表的第 27/28 列(Wind 导出区)。
    ws = wb["等权情绪指标"]
    close_rows = [[r[26], r[27]] for r in ws.iter_rows(min_row=3, values_only=True) if r[26] is not None]
    close = pd.DataFrame(close_rows, columns=["Date", "close"])
    close["Date"] = pd.to_datetime(close["Date"], errors="coerce")
    close["close"] = pd.to_numeric(close["close"], errors="coerce")
    close = close.dropna().set_index("Date").sort_index()

    base = close.join([pe, turn, illiq, fund], how="left").sort_index()
    base.index.name = "date"
    base = base.reset_index()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    base.to_csv(OUT, index=False)

    print({
        "csv": str(OUT.relative_to(ROOT)),
        "rows": int(len(base)),
        "range": [str(base["date"].min().date()), str(base["date"].max().date())],
        "non_null": {c: int(base[c].notna().sum()) for c in base.columns if c != "date"},
    })


if __name__ == "__main__":
    sys.exit(main())
